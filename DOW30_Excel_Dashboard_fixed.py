#!/usr/bin/env python3
"""
Revised DOW 30 Tracker (PyQt5 Edition)

This module improves upon the original DOW30_Excel_Dashboard script by fixing a
handful of issues that surfaced when the program was packaged as an
executable.  The main changes include:

* The `closeEvent` handler is now defined as a proper method of
  ``MainWindow`` rather than nested inside ``__init__``.  This prevents
  accidental indentation errors and ensures the application hides to the
  system tray instead of terminating when the user closes the window.
* Historical back‑fill scheduling now computes the correct date at run
  time.  Previously, the scheduler captured the current date when the
  program started, causing the application to overwrite yesterday’s sheet
  when it executed on subsequent days.  By calculating the historical
  timestamp on each invocation, a new workbook is created for the new
  day and the previous day’s data remains untouched.
* When running historical back‑fills we no longer perform a network
  request to Yahoo’s public API.  The fallback ``requests.get`` call is
  reserved exclusively for live fetches.  This reduces unnecessary API
  traffic before the scheduled time and avoids spurious blank or ``0``
  values.  For live fetches we still attempt to retrieve the
  ``regularMarketPrice`` using yfinance and only fall back to the
  unofficial quote endpoint when absolutely necessary.
* If no price can be retrieved for a ticker the cell is left blank.  We
  do not coerce missing data into ``0``, ensuring that the spreadsheet
  does not display misleading numbers.

Dependencies:
    pip install pandas requests schedule beautifulsoup4 openpyxl yfinance PyQt5

Build (one‑file EXE):
    pyinstaller --clean --onefile --windowed DOW30_Excel_Dashboard_fixed.py
"""

import os
import sys
import time
import threading
import traceback
from datetime import datetime, date, time as dtime, timedelta
from concurrent.futures import ThreadPoolExecutor

import schedule
import requests
import yfinance as yf
from PyQt5.QtCore import Qt, QTimer, pyqtSignal, QObject
from PyQt5.QtGui import QColor, QIcon
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QTableWidget, QTableWidgetItem,
    QToolBar, QAction, QCheckBox, QSystemTrayIcon, QMenu
)
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font as XLFont

# ── CONFIG ────────────────────────────────────────────────────────────────
TICKERS = [
    'AAPL','AMGN','AXP','BA','CAT','CRM','CSCO','CVX','DIS','DOW',
    'GS','HD','HON','IBM','INTC','JNJ','JPM','KO','MCD','MMM',
    'MRK','MSFT','NKE','PG','TRV','UNH','V','VZ','WBA','WMT'
]

# Define the hours we care about.  Keys are the 24‑hour hour, values are
# human‑friendly labels that appear as column headers in the spreadsheet.
HOURS = {
     9: "9:31 AM",   # one minute after market open
    10: "10:00 AM", 11: "11:00 AM", 12: "12 NOON",
    13: "1:00 PM",  14: "2:00 PM",  15: "3:00 PM", 16: "4:00 PM"
}

# Derive a stable path for saving workbooks next to this script.  Each
# workbook is named after the calendar date on which it is created.  The
# folder is created on demand.
BASE        = os.path.dirname(__file__)
SAVE_FOLDER = os.path.join(BASE, "Saved DOW Sheets")
os.makedirs(SAVE_FOLDER, exist_ok=True)

# ── EXCEL HELPERS ────────────────────────────────────────────────────────
def ensure_workbook() -> tuple[str, 'Workbook', 'Worksheet']:
    """Ensure that a workbook for today's date exists and return it.

    The workbook filename is derived from the current date in
    ``MM-DD-YYYY.xlsx`` format.  On the first call of the day the
    workbook and its header row are created; subsequent calls simply load
    the existing file.  The function always returns the path to the
    workbook, the ``Workbook`` object and the active worksheet titled
    ``"Prices"``.
    """
    fname = date.today().strftime("%m-%d-%Y") + ".xlsx"
    fpath = os.path.join(SAVE_FOLDER, fname)
    # Create a new workbook if one doesn't already exist for today
    if not os.path.exists(fpath):
        wb = Workbook()
        ws = wb.active
        ws.title = "Prices"
        # header row: ticker plus one column per hour label
        ws.append(["Ticker"] + list(HOURS.values()))
        # one row per ticker, initialise all values to None
        for t in TICKERS:
            ws.append([t] + [None] * len(HOURS))
        wb.save(fpath)
    wb = load_workbook(fpath)
    ws = wb["Prices"]
    return fpath, wb, ws


# ── FETCHER + SCHEDULER ──────────────────────────────────────────────────
class Fetcher(QObject):
    """Manage scheduled data collection and emit a signal upon updates."""

    updated = pyqtSignal()

    def __init__(self) -> None:
        super().__init__()
        # Always schedule creation of tomorrow's workbook.  If the program
        # happens to be running at midnight it will create the new file
        # before any data collection occurs.
        schedule.every().day.at("00:00").do(ensure_workbook)

        now = datetime.now()
        # Back‑fill any hours that have already passed today.  When the
        # program starts after market open we immediately populate the
        # spreadsheet with historical data for the current day.  We compute
        # ``hist_dt_today`` here to avoid capturing a stale date in the
        # lambda closures below.
        for hour, label in HOURS.items():
            hist_dt_today = datetime.combine(date.today(), dtime(hour=hour))
            if now >= hist_dt_today:
                # Immediate backfill for the current day
                self._fetch(label, use_history=True, hist_dt=hist_dt_today)
            # Schedule the same historical fetch for future days.  We pass
            # ``hour`` and ``label`` into the lambda and compute
            # ``hist_dt`` at run time using date.today().  Without this
            # computation, the original script would overwrite yesterday's
            # sheet by reusing the captured ``hist_dt``.
            schedule.every().day.at(f"{hour:02}:00").do(
                lambda l=label, h=hour: self._fetch(
                    l,
                    use_history=True,
                    hist_dt=datetime.combine(date.today(), dtime(hour=h))
                )
            )

        # Schedule future live quotes.  These tasks execute at the top of
        # each hour and fetch the current price.  The callback does not
        # provide a ``hist_dt`` which signals ``_fetch`` to treat the
        # operation as live.
        for hour, label in HOURS.items():
            schedule.every().day.at(f"{hour:02}:00").do(
                lambda l=label: self._fetch(l)
            )

        # Spin up a background thread to run the scheduler.  Using a
        # separate thread prevents the GUI event loop from blocking.
        threading.Thread(target=self._run_scheduler, daemon=True).start()

    def _run_scheduler(self) -> None:
        """Continuously run scheduled jobs in a dedicated thread."""
        while True:
            schedule.run_pending()
            time.sleep(1)

    def _fetch(self, label: str, use_history: bool = False, hist_dt: datetime | None = None) -> None:
        """Fetch prices for all tickers and write them to the workbook.

        Parameters
        ----------
        label: str
            The column header label corresponding to the hour being fetched.
        use_history: bool
            When ``True`` we fetch a narrow window of historical data
            surrounding ``hist_dt``.  When ``False`` we fetch the most
            recent market price.
        hist_dt: datetime | None
            The centre of the historical window.  This parameter must be
            provided when ``use_history`` is ``True``.  If ``None`` and
            ``use_history`` is ``True`` we compute the timestamp for
            ``label`` on today's date.
        """
        try:
            # Determine hist_dt on the fly if necessary
            if use_history and hist_dt is None:
                # Find the hour associated with the provided label.  This
                # reverse lookup allows us to compute the timestamp for
                # today's date.
                hour = next((h for h, lbl in HOURS.items() if lbl == label), None)
                if hour is None:
                    raise ValueError(f"Unknown label {label}")
                hist_dt = datetime.combine(date.today(), dtime(hour=hour))

            # Local helper to fetch a single ticker’s price.  Captures
            # ``use_history`` and ``hist_dt`` from the enclosing scope.
            def _get(sym: str) -> float | None:
                # If historical data is requested, query a two‑minute window
                # around ``hist_dt``.  Do not fall back to the public quote
                # endpoint during history fetches – if the historical data
                # isn’t available we return ``None`` instead of making an
                # unsolicited network request.
                if use_history and hist_dt is not None:
                    start = hist_dt - timedelta(minutes=1)
                    end   = hist_dt + timedelta(minutes=1)
                    df = yf.Ticker(sym).history(start=start, end=end, interval="1m")
                    if not df.empty:
                        # Use the last available close price in the window
                        return float(df["Close"].iloc[-1])
                    # If no data is returned we treat it as missing rather
                    # than querying the fallback endpoint.
                    return None

                # Live fetch: attempt to read regularMarketPrice from
                # yfinance, falling back to the unofficial Yahoo quote
                # endpoint if necessary.  This only executes when the
                # scheduled time arrives, thus avoiding premature GET
                # requests.
                info_price = yf.Ticker(sym).info.get("regularMarketPrice")
                if info_price is not None:
                    return float(info_price)
                # Only call the fallback endpoint if yfinance returns
                # nothing.  Wrap in try/except so that HTTP errors don’t
                # propagate to the scheduler thread.
                try:
                    resp = requests.get(
                        f"https://query1.finance.yahoo.com/v7/finance/quote?symbols={sym}",
                        timeout=10
                    )
                    resp.raise_for_status()
                    data = resp.json()
                    result = data.get("quoteResponse", {}).get("result", [])
                    if result:
                        p = result[0].get("regularMarketPrice")
                        return float(p) if p is not None else None
                except Exception:
                    # If the fallback fails we silently return None and
                    # allow the caller to leave the cell blank.
                    return None
                return None

            # Fetch all prices concurrently.  A modest thread pool improves
            # throughput without overwhelming the remote API.
            prices = list(ThreadPoolExecutor(max_workers=8).map(_get, TICKERS))

            # Open or create today's workbook
            fpath, wb, ws = ensure_workbook()
            # Determine the target column by looking up the header label
            try:
                col = list(HOURS.values()).index(label) + 2  # +2 because Excel is 1‑indexed and first col is ticker
            except ValueError:
                raise ValueError(f"Unrecognised column label {label}")

            # Define styling for up/down price movements
            up_fill   = PatternFill("solid", fgColor="C6EFCE")
            down_fill = PatternFill("solid", fgColor="FFC7CE")
            green     = XLFont(color="006100")
            red       = XLFont(color="9C0006")

            for row_idx, price in enumerate(prices, start=2):
                # Only write numeric values; leave missing data blank
                value = round(price, 2) if isinstance(price, (float, int)) else None
                cell  = ws.cell(row=row_idx, column=col, value=value)
                prev  = ws.cell(row=row_idx, column=col - 1).value
                if isinstance(prev, (float, int)) and isinstance(value, (float, int)):
                    # Colour the cell based on the direction of the change
                    if value > prev:
                        cell.fill, cell.font = up_fill, green
                    elif value < prev:
                        cell.fill, cell.font = down_fill, red

            wb.save(fpath)
            # Notify listeners that new data is available
            self.updated.emit()
        except Exception:
            # Log the traceback rather than crashing the scheduler thread
            traceback.print_exc()


# ── MAIN WINDOW ─────────────────────────────────────────────────────────
class MainWindow(QMainWindow):
    """Top‑level application window containing the table of prices."""

    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("DOW 30 Tracker")
        self.resize(1200, 700)
        # Create a system tray icon with a simple menu
        self.tray = QSystemTrayIcon(QIcon(), parent=self)
        tray_menu = QMenu(self)
        tray_menu.addAction("Show Window",   self._show_window)
        tray_menu.addAction("Refresh Now",    self.populate)
        tray_menu.addSeparator()
        tray_menu.addAction("Exit Tracker",   self._exit_app)
        self.tray.setContextMenu(tray_menu)
        self.tray.show()

        # Toolbar with toggles controlling how the table is displayed
        tb = self.addToolBar("Tools")
        tb.addAction("⟳ Refresh", self.populate)
        tb.addAction("Browse Excels…", lambda: os.startfile(SAVE_FOLDER) if os.name == 'nt' else os.system(f'xdg-open "{SAVE_FOLDER}"'))
        self.chkTimes = QCheckBox("Show Times")
        self.chkTimes.setChecked(True)
        self.chkPerc  = QCheckBox("Show % & Arrows")
        self.chkPerc.setChecked(True)
        self.chkStrip = QCheckBox("Stripe Rows")
        self.chkStrip.setChecked(True)
        for widget in (self.chkTimes, self.chkPerc, self.chkStrip):
            tb.addWidget(widget)

        # Table widget: first column for ticker symbols, remaining columns for times
        self.table = QTableWidget(len(TICKERS), len(HOURS) + 1, self)
        self.setCentralWidget(self.table)
        headers = ["Ticker"] + list(HOURS.values())
        self.table.setHorizontalHeaderLabels(headers)
        for idx, ticker in enumerate(TICKERS):
            item = QTableWidgetItem(ticker)
            item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
            self.table.setItem(idx, 0, item)

        # Start the fetcher and hook its update signal to our populate method
        self.fetcher = Fetcher()
        self.fetcher.updated.connect(self.populate)

        # Initial draw – delay slightly to allow the workbook to be created
        QTimer.singleShot(200, self.populate)

    def closeEvent(self, event) -> None:
        """Override the window close event to minimise to the system tray."""
        event.ignore()
        self.hide()
        # Notify the user that the application is still running in the background
        self.tray.showMessage(
            "DOW 30 Tracker",
            "Still running in background. Double‑click tray icon to restore.",
            QSystemTrayIcon.Information,
            2000
        )

    def _show_window(self) -> None:
        """Restore the main window from the system tray."""
        self.show()
        self.activateWindow()

    def _exit_app(self) -> None:
        """Quit the application cleanly."""
        QApplication.quit()

    def populate(self) -> None:
        """Redraw the table using the latest data from today's workbook."""
        fpath, wb, ws = ensure_workbook()
        # Skip the header row when reading values
        rows = list(ws.values)[1:]
        for r, row in enumerate(rows):
            prev = None
            for c, raw in enumerate(row):
                item = QTableWidgetItem()
                text = ""
                colour = None
                if c == 0:
                    # Ticker column – show the symbol
                    text = raw
                else:
                    if not self.chkTimes.isChecked():
                        text = ""
                    else:
                        if isinstance(raw, (int, float)):
                            if self.chkPerc.isChecked() and prev is not None and prev:
                                diff  = raw - prev
                                pct   = (diff / prev * 100) if prev != 0 else 0
                                arrow = "▲" if diff > 0 else ("▼" if diff < 0 else "")
                                text  = f"{arrow}{raw:.2f} ({pct:+.2f}%)"
                                colour = QColor("#006100") if diff > 0 else QColor("#9C0006") if diff < 0 else None
                            else:
                                text = f"{raw:.2f}"
                        else:
                            text = ""
                    # Track the previous numeric value for percentage calculations
                    prev = raw if isinstance(raw, (int, float)) else prev
                item.setText(text)
                if colour:
                    item.setForeground(colour)
                item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
                self.table.setItem(r, c, item)
            # Optional row striping for readability
            if self.chkStrip.isChecked():
                bg_colour = QColor("#F7F7F7") if (r % 2) else QColor("#FFFFFF")
                for c in range(self.table.columnCount()):
                    self.table.item(r, c).setBackground(bg_colour)
        # Adjust column widths based on content
        self.table.resizeColumnsToContents()


# ── RUN ───────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    # Start the Qt event loop
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())