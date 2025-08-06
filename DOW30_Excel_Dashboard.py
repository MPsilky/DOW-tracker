#!/usr/bin/env python3
"""
DOW 30 Tracker (PyQt5 Edition)

Dependencies:
    pip install pandas requests schedule beautifulsoup4 openpyxl yfinance PyQt5

Build (one‑file EXE):
    pyinstaller --clean --onefile --windowed DOW30_Excel_Dashboard.py
"""
import os
import sys
import time                                  # ← needed for scheduler loop
import threading
import traceback
import schedule
import requests
from datetime import datetime, date, time as dtime, timedelta
from concurrent.futures import ThreadPoolExecutor

import yfinance as yf
from PyQt5.QtCore import Qt, QTimer, pyqtSignal, QObject
from PyQt5.QtGui import QColor, QIcon
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QTableWidget, QTableWidgetItem,
    QToolBar, QAction, QCheckBox, QSystemTrayIcon, QMenu
)

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font as XLFont

# ── CONFIG ────────────────────────────────────────────────────────────────
TICKERS = [
    'AAPL','AMGN','AXP','BA','CAT','CRM','CSCO','CVX','DIS','DOW',
    'GS','HD','HON','IBM','INTC','JNJ','JPM','KO','MCD','MMM',
    'MRK','MSFT','NKE','PG','TRV','UNH','V','VZ','WBA','WMT'
]
HOURS = {
     9: "9:31 AM",   # as requested +1 min after market open
    10: "10:00 AM", 11: "11:00 AM", 12: "12 NOON",
    13: "1:00 PM",  14: "2:00 PM",  15: "3:00 PM", 16: "4:00 PM"
}
BASE        = os.path.dirname(__file__)
SAVE_FOLDER = os.path.join(BASE, "Saved DOW Sheets")
os.makedirs(SAVE_FOLDER, exist_ok=True)

# ── EXCEL HELPERS ────────────────────────────────────────────────────────
def ensure_workbook():
    fname = date.today().strftime("%m-%d-%Y") + ".xlsx"
    fpath = os.path.join(SAVE_FOLDER, fname)
    if not os.path.exists(fpath):
        wb = Workbook()
        ws = wb.active
        ws.title = "Prices"
        ws.append(["Ticker"] + list(HOURS.values()))
        for t in TICKERS:
            ws.append([t] + [None]*len(HOURS))
        wb.save(fpath)
    wb = load_workbook(fpath)
    ws = wb["Prices"]
    return fpath, wb, ws

# ── FETCHER + SCHEDULER ──────────────────────────────────────────────────
class Fetcher(QObject):
    updated = pyqtSignal()

    def __init__(self):
        super().__init__()

        # 1) ensure tomorrow’s sheet gets created at midnight
        schedule.every().day.at("00:00").do(ensure_workbook)

        now = datetime.now()
        # 2) back‑fill any past hours right away, then still schedule them
        for h, lbl in HOURS.items():
            hist_dt = datetime.combine(now.date(), dtime(hour=h))
            if now >= hist_dt:
                # immediate backfill
                self._fetch(lbl, use_history=True, hist_dt=hist_dt)
            # schedule that same historical fetch for future days
            schedule.every().day.at(f"{h:02}:00")\
                    .do(lambda l=lbl, d=hist_dt: self._fetch(l, True, d))

        # 3) schedule future live quotes (no history flag)
        for h, lbl in HOURS.items():
            schedule.every().day.at(f"{h:02}:00")\
                    .do(lambda l=lbl: self._fetch(l))

        # 4) run the scheduler loop
        threading.Thread(target=self._run, daemon=True).start()

    def _run(self):
        while True:
            schedule.run_pending()
            time.sleep(1)

    def _fetch(self, label, use_history=False, hist_dt=None):
        try:
            def _get(sym):
                # historical pull if requested
                if use_history and hist_dt:
                    start = hist_dt - timedelta(minutes=1)
                    end   = hist_dt + timedelta(minutes=1)
                    df = yf.Ticker(sym).history(
                        start=start, end=end, interval="1m"
                    )
                    if not df.empty:
                        return float(df["Close"].iloc[-1])
                # live fallback
                p = yf.Ticker(sym).info.get("regularMarketPrice")
                if p is None:
                    j = requests.get(
                        f"https://query1.finance.yahoo.com/v7/finance/quote?symbols={sym}"
                    ).json()
                    p = j["quoteResponse"]["result"][0].get("regularMarketPrice")
                return float(p) if p else None

            prices = list(ThreadPoolExecutor(8).map(_get, TICKERS))
            fpath, wb, ws = ensure_workbook()
            col = list(HOURS.values()).index(label) + 2

            up_fill   = PatternFill("solid", fgColor="C6EFCE")
            down_fill = PatternFill("solid", fgColor="FFC7CE")
            green     = XLFont(color="006100")
            red       = XLFont(color="9C0006")

            for r, price in enumerate(prices, start=2):
                val  = round(price, 2) if isinstance(price, (float,int)) else None
                cell = ws.cell(row=r, column=col, value=val)
                prev = ws.cell(row=r, column=col-1).value
                if isinstance(prev, (float,int)) and isinstance(val, (float,int)):
                    if val > prev:
                        cell.fill, cell.font = up_fill, green
                    elif val < prev:
                        cell.fill, cell.font = down_fill, red

            wb.save(fpath)
            self.updated.emit()

        except Exception:
            traceback.print_exc()

# ── MAIN WINDOW ─────────────────────────────────────────────────────────
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("DOW 30 Tracker")
        self.resize(1200, 700)
        self.tray = QSystemTrayIcon(QIcon(), parent=self)
        tray_menu = QMenu(self)
        tray_menu.addAction("Show Window",   self._show_window)
        tray_menu.addAction("Refresh Now",    self.populate)
        tray_menu.addSeparator()
        tray_menu.addAction("Exit Tracker",   self._exit_app)
        self.tray.setContextMenu(tray_menu)
        self.tray.show()

        def closeEvent(self, event):
        # Hide window instead of quitting
          event.ignore()
        self.hide()
        self.tray.showMessage(
            "DOW 30 Tracker",
            "Still running in background. Double‑click tray icon to restore.",
            QSystemTrayIcon.Information,
            2000
        )

    def _show_window(self):
        self.show()
        self.activateWindow()

    def _exit_app(self):
        QApplication.quit()

        # toolbar
        tb = self.addToolBar("Tools")
        tb.addAction("⟳ Refresh", self.populate)
        tb.addAction("Browse Excels…", lambda: os.startfile(SAVE_FOLDER))
        self.chkTimes = QCheckBox("Show Times")
        self.chkTimes.setChecked(True)
        self.chkPerc  = QCheckBox("Show % & Arrows")
        self.chkPerc.setChecked(True)
        self.chkStrip = QCheckBox("Stripe Rows")
        self.chkStrip.setChecked(True)
        for w in (self.chkTimes, self.chkPerc, self.chkStrip):
            tb.addWidget(w)

        # table
        self.table = QTableWidget(len(TICKERS), len(HOURS)+1, self)
        self.setCentralWidget(self.table)
        headers = ["Ticker"] + list(HOURS.values())
        self.table.setHorizontalHeaderLabels(headers)
        for i, t in enumerate(TICKERS):
            item = QTableWidgetItem(t)
            item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
            self.table.setItem(i, 0, item)

        # start fetcher & hook its signal
        self.fetcher = Fetcher()
        self.fetcher.updated.connect(self.populate)

        # initial draw (a little later so workbook is ready)
        QTimer.singleShot(200, self.populate)

    def populate(self):
        fpath, wb, ws = ensure_workbook()
        data = list(ws.values)[1:]  # skip header
        for r, row in enumerate(data):
            prev = None
            for c, raw in enumerate(row):
                item = QTableWidgetItem()
                text = ""
                color = None

                if c == 0:
                    text = raw
                else:
                    if not self.chkTimes.isChecked():
                        text = ""
                    else:
                        if isinstance(raw, (int, float)):
                            if self.chkPerc.isChecked() and prev is not None:
                                diff  = raw - prev
                                pct   = (diff/prev*100) if prev else 0
                                arrow = "▲" if diff>0 else ("▼" if diff<0 else "")
                                text  = f"{arrow}{raw:.2f} ({pct:+.2f}%)"
                                color = QColor("#006100") if diff>0 else QColor("#9C0006") if diff<0 else None
                            else:
                                text = f"{raw:.2f}"
                        else:
                            text = ""
                    prev = raw if isinstance(raw, (int,float)) else prev

                item.setText(text)
                if color:
                    item.setForeground(color)
                item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
                self.table.setItem(r, c, item)

            # row striping
            if self.chkStrip.isChecked():
                bg = QColor("#F7F7F7") if (r % 2) else QColor("#FFFFFF")
                for c in range(self.table.columnCount()):
                    self.table.item(r, c).setBackground(bg)

        self.table.resizeColumnsToContents()


# ── RUN ───────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = QApplication(sys.argv)
    win = MainWindow()
    win.show()
    sys.exit(app.exec_())
